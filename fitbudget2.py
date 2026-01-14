import streamlit as st
import numpy as np
import pandas as pd
import unicodedata
import time
from functools import reduce
from io import BytesIO

# ＊스타일 구역＊
st.markdown(
    """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Nanum+Gothic+Coding&display=swap');
        .stTextInput, .stButton > button, .stSelectbox, .stDateInput, .stTimeInput, 
        input[type="number"], code[class="language-java"], p, input[type="text"],
        textarea[aria-label="결과 출력"]{
            font-family: 'Nanum Gothic Coding', monospace !important;
            font-size: 14px;color: #FFC83D;}
        input[type="number"] { text-align: right; }
        h1{ text-align: center;}        
        h3{ text-align: right; margin-right: 0;margin-top: 0;padding-top: 0;padding-right: 0;line-height: 1.2;}        
        [data-testid="stCheckbox"] {
            margin-left: 5px;
            margin-right: -5px;
            height: 1rem;
            width: 1rem;}
        input[type="number"], textarea[aria-label="결과 출력"], input[type="text"], 
        [data-testid="stVerticalBlock"] > div:first-child {margin: 2px;}
        input[aria-label="budget"]{margin: 0px;font-size: 24px;font-weight: bold;}
        [data-testid="stNotificationContentWarning"]{margin: -8px; padding: 6px 12px !important; font-size: 16px;}
        [data-testid="stNotificationContentWarning"] p {margin: 0 !important; padding: 0 !important;}
        [data-testid="stAlert"]{padding: 6px 8px !important; margin: 0 !important;}
        [data-testid="stHorizontalBlock"] { margin-bottom: -18px; }
        [data-testid="column"] { margin-right: -4px; margin-left: -4px; }
        .stDataFrame { width: 100% !important; }
        h3, p { color: #FFC83D; }
        [data-testid="baseButton-secondary"],[data-testid="stDataFrameResizable"]{width: 100% !important;}
        
        /* 물품추가, 계산하기 버튼 스타일 (primary 버튼) */
        [data-testid="stBaseButton-primary"] {
            border: 2px solid #FFC83D !important;
            border-radius: 12px !important;
            background-color: transparent !important;
            color: #FFC83D !important;
            padding: 6px 6px !important;
            margin-top: 12px 0 !important;
            font-weight: bold !important;
            transition: all 0.3s ease !important;
            min-height: auto !important;
            line-height: 1.4 !important;
            white-space: nowrap !important;
        }
        [data-testid="stBaseButton-primary"]:hover,
        [data-testid="stBaseButton-primary"]:focus,
        [data-testid="stBaseButton-primary"]:active {
            background-color: #FFC83D !important;
            color: #000000 !important;
        }
        [data-testid="stBaseButton-primary"] p {
            color: inherit !important;
        }
        
        /* 정렬 버튼 스타일 (기존 유지 - 일반 버튼 스타일 오버라이드) */
        button[kind="secondary"][data-testid="stBaseButton-secondary"] {
            background-color: transparent !important;
            border: none !important;
            border-radius: 0 !important;
            padding: 0 !important;
            font-size: 14px !important;
            color: #FFC83D !important;
            cursor: pointer;
        }
        button[kind="secondary"][data-testid="stBaseButton-secondary"]:hover {
            background-color: transparent !important;
            color: #FFE082 !important;
            text-decoration: underline;
        }
    </style>""", unsafe_allow_html=True)

# ＊함수 구역＊
def get_print_length(s):
    screen_length = 0
    for char in s:
        if unicodedata.east_asian_width(char) in ['F', 'W']:
            screen_length += 2
        else:
            screen_length += 1
    return screen_length

def cut_string(org_s, max_length, pad_LR="R"):
    cut_s, length = '', 0
    for char in org_s:
        char_length = get_print_length(char)
        if length + char_length > max_length:
            break
        cut_s += char
        length += char_length
    diff = max_length - length
    if diff > 0:
        if pad_LR == "L":
            return diff * " " + cut_s
        if pad_LR == "R":
            return cut_s + diff * " "
    else:
        return cut_s

def update_all_items():
    """예산 변경 시 모든 아이템의 최소/최대 구매 가능 개수 업데이트"""
    budget = st.session_state.get("budget", 0)
    if budget <= 0:
        return
    
    for i in range(st.session_state.get('item_count', 5)):
        item_price = st.session_state.get(f"item_price_{i}", 0)
        if item_price > 0 and item_price <= budget:
            max_possible = budget // item_price
            # 최대 구매 가능 개수 상한 설정
            st.session_state[f"item_max_limit_{i}"] = max_possible
            st.session_state[f"item_disabled_{i}"] = False
            
            # 최대구매: 0이면 max_possible로 설정, 상한 초과하면 조정
            current_max = st.session_state.get(f"item_max_{i}", 0)
            if current_max == 0 or current_max > max_possible:
                st.session_state[f"item_max_{i}"] = max_possible
            
            # 최소구매: 최대값 초과하면 조정 (0 초기화는 유지)
            current_min = st.session_state.get(f"item_min_{i}", 0)
            if current_min > st.session_state.get(f"item_max_{i}", 0):
                st.session_state[f"item_min_{i}"] = st.session_state.get(f"item_max_{i}", 0)
        else:
            st.session_state[f"item_disabled_{i}"] = True
            st.session_state[f"item_max_{i}"] = 0

def on_budget_change():
    """예산 변경 콜백"""
    update_all_items()

def on_price_change(i):
    """단가 변경 콜백"""
    budget = st.session_state.get("budget", 0)
    item_price = st.session_state.get(f"item_price_{i}", 0)
    
    if budget > 0 and item_price > 0 and item_price <= budget:
        max_possible = budget // item_price
        st.session_state[f"item_max_limit_{i}"] = max_possible
        st.session_state[f"item_disabled_{i}"] = False
        
        # 최대값 초기화 또는 조정
        current_max = st.session_state.get(f"item_max_{i}", 0)
        if current_max == 0 or current_max > max_possible:
            st.session_state[f"item_max_{i}"] = max_possible
    else:
        st.session_state[f"item_disabled_{i}"] = True

def on_min_change(i):
    """최소 구매량 변경 콜백"""
    current_min = st.session_state.get(f'item_min_{i}', 0)
    current_max = st.session_state.get(f'item_max_{i}', 0)
    
    if current_min > current_max:
        st.session_state[f'item_min_{i}'] = current_max

def on_max_change(i):
    """최대 구매량 변경 콜백"""
    current_max = st.session_state.get(f"item_max_{i}", 0)
    current_min = st.session_state.get(f'item_min_{i}', 0)
    current_price = st.session_state.get(f'item_price_{i}', 0)
    budget = st.session_state.get("budget", 0)
    
    if current_price > 0 and (current_price * current_max) > budget:
        st.session_state[f'item_max_{i}'] = budget // current_price
    elif current_min > current_max:
        st.session_state[f'item_max_{i}'] = current_min

def calculate_budget(budget, labels, prices, base_quantity, limited_quantity):
    """메모이제이션 + 마지막 아이템 나눗셈 처리"""
    try:
        text_out = f'사용해야 할 예산은 {format(budget,",")}원입니다.\n'
        item_count = len(prices)
        
        combined = zip(prices, labels, base_quantity, limited_quantity)
        sorted_combined = sorted(combined, reverse=True)
        prices, labels, base_quantity, limited_quantity = map(list, zip(*sorted_combined))
        
        text_width = 25
        text_out += '_' * text_width + '정렬된 데이터' + '_' * text_width + '\n'
        for n_prt in range(item_count):
            label = cut_string(labels[n_prt], 28)
            text_out += f"품목 #{n_prt + 1:02d} {label} = {prices[n_prt]:7,d} 원 ({base_quantity[n_prt]:3d}  ~ {limited_quantity[n_prt]:3d})\n"
        text_out += '_' * (text_width * 2 + 13) + '\n'
        
        total_budget = budget
        fixed_budget = sum(a * b for a, b in zip(base_quantity, prices))
        remaining_budget = budget - fixed_budget
        limits = [lim - base for lim, base in zip(limited_quantity, base_quantity)]
        last_idx = item_count - 1
        
        time_limit = 20
        start_time = time.time()
        memo = {}
        call_count = 0
        
        def count_solutions(idx, remaining):
            nonlocal call_count
            call_count += 1
            
            if call_count % 100000 == 0:
                if time.time() - start_time > time_limit:
                    raise TimeoutError(f"시간초과: {time_limit}초 경과")
            
            if remaining < 0:
                return 0
            
            if idx == last_idx:
                qty = remaining // prices[last_idx]
                if qty <= limits[last_idx] and remaining % prices[last_idx] == 0:
                    return 1
                return 0
            
            if (idx, remaining) in memo:
                return memo[(idx, remaining)]
            
            total = 0
            for qty in range(limits[idx] + 1):
                cost = qty * prices[idx]
                if cost > remaining:
                    break
                total += count_solutions(idx + 1, remaining - cost)
            
            memo[(idx, remaining)] = total
            return total
        
        exact_count = count_solutions(0, remaining_budget)
        
        cases_exact = []
        
        def reconstruct(idx, remaining, current):
            if time.time() - start_time > time_limit:
                raise TimeoutError(f"시간초과: {time_limit}초 경과")
            
            if idx == last_idx:
                if remaining % prices[last_idx] == 0:
                    qty = remaining // prices[last_idx]
                    if qty <= limits[last_idx]:
                        cases_exact.append(current + [qty])
                return
            
            for qty in range(limits[idx] + 1):
                cost = qty * prices[idx]
                if cost > remaining:
                    break
                
                next_remaining = remaining - cost
                if memo.get((idx + 1, next_remaining), 0) > 0:
                    current.append(qty)
                    reconstruct(idx + 1, next_remaining, current)
                    current.pop()
                elif idx + 1 == last_idx:
                    if next_remaining % prices[last_idx] == 0:
                        qty_last = next_remaining // prices[last_idx]
                        if qty_last <= limits[last_idx]:
                            cases_exact.append(current + [qty, qty_last])
        
        if exact_count > 0:
            reconstruct(0, remaining_budget, [])
        
        cases_close = []
        if exact_count == 0:
            best_remaining = remaining_budget
            
            def find_closest(idx, remaining, current):
                nonlocal best_remaining
                if time.time() - start_time > time_limit:
                    return
                
                if idx == last_idx:
                    qty = min(remaining // prices[last_idx], limits[last_idx])
                    leftover = remaining - qty * prices[last_idx]
                    if leftover < best_remaining:
                        best_remaining = leftover
                        cases_close.clear()
                        cases_close.append(current + [qty])
                    elif leftover == best_remaining:
                        cases_close.append(current + [qty])
                    return
                
                for qty in range(limits[idx] + 1):
                    cost = qty * prices[idx]
                    if cost > remaining:
                        break
                    current.append(qty)
                    find_closest(idx + 1, remaining - cost, current)
                    current.pop()
            
            find_closest(0, remaining_budget, [])
        
        end_time = time.time()
        execution_time = end_time - start_time
        print(f"실행 시간: {execution_time:.4f}초, 메모 상태 수: {len(memo):,}")
        
        if exact_count == 0:
            text_out += f'{total_budget:,d}원의 예산에 맞게 구입할 방법이 없습니다.\n'
            text_out += '예산에 근접한 구입 계획은 아래와 같습니다.\n'
            list_show = cases_close
        else:
            text_out += f'예산에 맞는 {len(cases_exact):,d}개의 완벽한 방법을 찾았습니다.\n'
            list_show = cases_exact
        
        list_show = (np.array(list_show) + np.array(base_quantity)).tolist() if list_show else []
        text_out += f'이 프로그램은 {call_count:,d}개의 상태를 계산했습니다.\n'
        
        # labels도 함께 반환
        return text_out, list_show, prices, labels
    
    except TimeoutError as e:
        return f'에러입니다.: {e}', [], prices, labels
    except Exception as e:
        print('Error Message:', e)
        return f'에러입니다.: {e}', [], prices, labels

def create_template_excel():
    """엑셀 양식 생성 (단일 시트)"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 예산 행 + 물품 목록을 하나의 시트에
        rows = [
            ['예산', 100000, '', ''],
            ['', '', '', ''],
            ['물품이름', '단가', '최소구매', '최대구매'],
            ['물품1', 10000, 0, 10],
            ['물품2', 15000, 0, 6],
            ['물품3', 20000, 0, 5],
        ]
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='예산계산', index=False, header=False)
    
    output.seek(0)
    return output

def extract_number(value):
    """문자열에서 숫자만 추출 (원, 개 등 단위 제거)"""
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return value
    # 문자열에서 숫자와 소수점만 추출
    import re
    nums = re.sub(r'[^\d.]', '', str(value))
    if nums:
        return float(nums)
    return None

def load_from_excel(uploaded_file):
    """엑셀 파일에서 데이터 로드 (단일 시트)"""
    try:
        df = pd.read_excel(uploaded_file, sheet_name=0, header=None)
        
        # 첫 행에서 예산 읽기 (A1='예산', B1=값)
        budget = int(extract_number(df.iloc[0, 1]))
        
        # 3행부터 물품 데이터 (3행은 헤더: 물품이름, 단가, 최소구매, 최대구매)
        df_items = df.iloc[3:].copy()
        df_items.columns = ['물품이름', '단가', '최소구매', '최대구매']
        
        # 단가: 단위 제거 후 숫자 변환
        df_items['단가'] = df_items['단가'].apply(extract_number)
        df_items['물품이름'] = df_items['물품이름'].fillna('')
        
        # 단가가 없거나 0인 행 제외
        df_items = df_items.dropna(subset=['단가'])
        df_items = df_items[df_items['단가'] > 0]
        df_items = df_items.reset_index(drop=True)
        
        # 최소구매, 최대구매: 숫자 변환 후 예산 기준 검증
        items_data = []
        for _, row in df_items.iterrows():
            price = int(row['단가'])
            max_possible = budget // price if price > 0 else 0
            
            # 최소구매: NaN이면 0, 불가능하면 0
            min_qty = extract_number(row['최소구매'])
            if min_qty is None or min_qty < 0 or min_qty > max_possible:
                min_qty = 0
            else:
                min_qty = int(min_qty)
            
            # 최대구매: NaN이면 max_possible, 불가능하면 max_possible
            max_qty = extract_number(row['최대구매'])
            if max_qty is None or max_qty <= 0 or max_qty > max_possible:
                max_qty = max_possible
            else:
                max_qty = int(max_qty)
            
            # 최소가 최대보다 크면 최소를 0으로
            if min_qty > max_qty:
                min_qty = 0
            
            items_data.append({
                '물품이름': row['물품이름'],
                '단가': price,
                '최소구매': min_qty,
                '최대구매': max_qty
            })
        
        df_result = pd.DataFrame(items_data)
        return budget, df_result
    except Exception as e:
        st.error(f"파일 로드 오류: {e}")
        return None, None

def create_result_excel(result_text, df_result, result_labels=None):
    """결과를 엑셀 파일로 생성 (단일 시트) - 품목 이름 행 추가, 필터 및 셀 병합"""
    from openpyxl.utils import get_column_letter
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 텍스트 결과를 행으로 변환
        text_lines = result_text.split('\n')
        rows = [[line] for line in text_lines]
        
        # 결과요약 행 수 저장 (셀 병합용)
        summary_row_count = len(rows)
        
        # 빈 행 추가
        rows.append([''])
        summary_row_count += 1
        
        # 가격 헤더 행 번호 저장 (필터용)
        price_header_row = None
        
        # DataFrame 헤더 추가
        if df_result is not None and len(df_result) > 0:
            # 품목 번호 행 추가 (#01, #02, ... 형식, 금액 컬럼은 빈 문자열)
            if result_labels:
                num_row = [f'#{i+1:02d}' for i in range(len(result_labels))] + ['']
                rows.append(num_row)
            
            # 품목 이름 행 추가 (금액 컬럼은 빈 문자열)
            if result_labels:
                name_row = result_labels + ['']
                rows.append(name_row)
            
            # 가격 헤더 행
            price_header_row = len(rows) + 1  # 1-based index for Excel
            rows.append(df_result.columns.tolist())
            
            # DataFrame 데이터 추가
            for _, row in df_result.iterrows():
                rows.append(row.tolist())
        
        df_output = pd.DataFrame(rows)
        df_output.to_excel(writer, sheet_name='계산결과', index=False, header=False)
        
        # openpyxl로 추가 작업
        ws = writer.sheets['계산결과']
        
        # 1. 상단 결과요약 A-G열 행별 셀 병합
        for row_idx in range(1, summary_row_count + 1):
            ws.merge_cells(f'A{row_idx}:G{row_idx}')
        
        # 2. 가격 헤더 행에 필터 적용
        if price_header_row and df_result is not None and len(df_result) > 0:
            num_cols = len(df_result.columns)
            last_col = get_column_letter(num_cols)
            last_row = price_header_row + len(df_result)
            ws.auto_filter.ref = f'A{price_header_row}:{last_col}{last_row}'
    
    output.seek(0)
    return output

# ＊메인 UI＊
result_text = '''예산과 단가를 입력한 후\n계산하기 버튼을 누르면,
예산에 딱 맞게 물건을\n살 수 있는 방법을 찾아줍니다.\n
데이터프레임으로 출력된 결과에
마우스를 올리면 저장도 가능해요.\n
물품 이름은 안 쓰셔도 작동합니다.
단가가 0인 품목은 자동으로 제외합니다.
물품 추가 버튼을 눌러\n물품을 추가할 수도 있고,
체크 박스의 체크 표시를 해제하면\n잠시 계산에서 제외할 수도 있습니다.
기본 구매량과 최대 구매량을 제한할 수 있습니다.
'''

result_list, result_prices, result_labels = [], [], []  # result_labels 추가

st.title("편리한 예산🍞만들기")
st.markdown('<p style="color: #a8a888;text-align: right;">SimBud beta (Budget Simulator V2.00)by 교사 박현수, 버그 및 개선 문의: <a href="mailto:hanzch84@gmail.com">hanzch84@gmail.com</a></p>', unsafe_allow_html=True)

# 엑셀 업로드/다운로드 섹션
with st.expander("📁 엑셀 파일로 관리하기", expanded=False):
    col_download, col_upload = st.columns(2)
    
    with col_download:
        st.write("**양식 다운로드**")
        template_excel = create_template_excel()
        st.download_button(
            label="📥 양식 다운로드",
            data=template_excel,
            file_name="예산계산_양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col_upload:
        st.write("**파일 업로드**")
        uploaded_file = st.file_uploader("엑셀 파일 선택", type=['xlsx'], label_visibility='collapsed', key='excel_uploader')
        
        if uploaded_file is not None:
            # 이미 로드한 파일인지 확인 (파일명과 크기로 체크)
            file_key = f"{uploaded_file.name}_{uploaded_file.size}"
            if st.session_state.get('loaded_file_key') != file_key:
                budget_loaded, df_items_loaded = load_from_excel(uploaded_file)
                if budget_loaded is not None:
                    # 세션 스테이트 초기화 및 데이터 로드
                    st.session_state['budget'] = budget_loaded
                    # 기본 5줄, 업로드된 개수가 더 많으면 그만큼 생성
                    st.session_state['item_count'] = max(5, len(df_items_loaded))
                    
                    for i, (_, row) in enumerate(df_items_loaded.iterrows()):
                        st.session_state[f'item_name_{i}'] = str(row['물품이름']) if pd.notna(row['물품이름']) else ''
                        st.session_state[f'item_price_{i}'] = int(row['단가'])
                        st.session_state[f'item_min_{i}'] = int(row['최소구매'])
                        st.session_state[f'item_max_{i}'] = int(row['최대구매'])
                        st.session_state[f'item_usable_{i}'] = True
                    
                    update_all_items()
                    st.session_state['loaded_file_key'] = file_key
                    st.success(f"✅ 데이터 로드 완료! 예산: {budget_loaded:,}원, 물품: {len(df_items_loaded)}개")
                    st.rerun()
            else:
                st.success(f"✅ 데이터 로드 완료! 예산: {st.session_state.get('budget', 0):,}원, 물품: {st.session_state.get('item_count', 0)}개")

# 예산 입력
col_label_budget, col_input_budget = st.columns([3, 7])
with col_label_budget:
    st.subheader("사용할 예산")
with col_input_budget:
    budget_input = st.number_input(
        "budget", 
        min_value=0, 
        key="budget", 
        help="사용해야하는 예산을 입력하세요.",
        on_change=on_budget_change, 
        format="%d", 
        label_visibility='collapsed'
    )

# session_state 초기화
if 'item_count' not in st.session_state:
    st.session_state.item_count = 5

# 정렬 상태 초기화
if 'sort_key' not in st.session_state:
    st.session_state.sort_key = None
    st.session_state.sort_ascending = True

def apply_sort():
    """정렬 적용 - session_state 데이터를 실제로 재정렬"""
    sort_key = st.session_state.sort_key
    if sort_key is None:
        return
    
    # 현재 데이터 수집
    items = []
    for i in range(st.session_state.item_count):
        items.append({
            'name': st.session_state.get(f'item_name_{i}', ''),
            'price': st.session_state.get(f'item_price_{i}', 0),
            'min': st.session_state.get(f'item_min_{i}', 0),
            'max': st.session_state.get(f'item_max_{i}', 0),
            'usable': st.session_state.get(f'item_usable_{i}', True),
            'disabled': st.session_state.get(f'item_disabled_{i}', True),
            'max_limit': st.session_state.get(f'item_max_limit_{i}', 9999),
        })
    
    ascending = st.session_state.sort_ascending
    
    if sort_key == 'name':
        items.sort(key=lambda x: x['name'], reverse=not ascending)
    elif sort_key == 'min':
        items.sort(key=lambda x: x['min'], reverse=not ascending)
    elif sort_key == 'max':
        items.sort(key=lambda x: x['max'], reverse=not ascending)
    elif sort_key == 'price':
        items.sort(key=lambda x: x['price'], reverse=not ascending)
    
    # 정렬된 순서로 session_state 업데이트
    for i, item in enumerate(items):
        st.session_state[f'item_name_{i}'] = item['name']
        st.session_state[f'item_price_{i}'] = item['price']
        st.session_state[f'item_min_{i}'] = item['min']
        st.session_state[f'item_max_{i}'] = item['max']
        st.session_state[f'item_usable_{i}'] = item['usable']
        st.session_state[f'item_disabled_{i}'] = item['disabled']
        st.session_state[f'item_max_limit_{i}'] = item['max_limit']

def toggle_sort(key):
    """정렬 토글 함수"""
    if st.session_state.sort_key == key:
        st.session_state.sort_ascending = not st.session_state.sort_ascending
    else:
        st.session_state.sort_key = key
        st.session_state.sort_ascending = True
    apply_sort()

def get_sort_indicator(key):
    """정렬 방향 표시"""
    if st.session_state.sort_key == key:
        return " ▲" if st.session_state.sort_ascending else " ▼"
    return ""

# 아이템 헤더 (정렬 버튼)
hcol1, hcol2, hcol3, hcol4, hcol5 = st.columns([3.5, 1.4, 1.4, 3, 0.7])
with hcol1:
    if st.button(f"물품이름{get_sort_indicator('name')}", key="sort_name"):
        toggle_sort('name')
        st.rerun()
with hcol2:
    if st.button(f"최소{get_sort_indicator('min')}", key="sort_min"):
        toggle_sort('min')
        st.rerun()
with hcol3:
    if st.button(f"최대{get_sort_indicator('max')}", key="sort_max"):
        toggle_sort('max')
        st.rerun()
with hcol4:
    if st.button(f"물품단가{get_sort_indicator('price')}", key="sort_price"):
        toggle_sort('price')
        st.rerun()
with hcol5:
    st.write("선택")

# 아이템 입력 필드 생성
item_names = []
item_prices = []
min_quantities = []
max_quantities = []

for i in range(st.session_state.item_count):
    col1, col2, col3, col4, col5 = st.columns([3.5, 1.4, 1.4, 3, 0.7])
    is_disabled = not st.session_state.get(f'item_usable_{i}', True)
    
    with col1:
        item_name = st.text_input(
            f"물품{i+1} 이름 입력", 
            label_visibility='collapsed',
            key=f"item_name_{i}", 
            placeholder=f"물품{i+1} 이름 입력",
            disabled=is_disabled
        )
    
    with col2:
        max_limit = st.session_state.get(f'item_max_limit_{i}', 9999)
        item_min = st.number_input(
            f"최소 {i+1}",
            min_value=0,
            max_value=max_limit,
            key=f"item_min_{i}",
            on_change=on_min_change,
            args=(i,),
            disabled=is_disabled or st.session_state.get(f"item_disabled_{i}", True),
            format="%d", 
            label_visibility='collapsed'
        )
    
    with col3:
        item_max = st.number_input(
            f"최대 {i+1}",
            min_value=0,
            max_value=max_limit,
            key=f"item_max_{i}",
            on_change=on_max_change,
            args=(i,),
            disabled=is_disabled or st.session_state.get(f"item_disabled_{i}", True),
            format="%d", 
            label_visibility='collapsed'
        )
    
    with col4:
        item_price = st.number_input(
            f"물품단가{i+1}",
            min_value=0,
            key=f"item_price_{i}",
            on_change=on_price_change,
            args=(i,),
            disabled=is_disabled, 
            format="%d", 
            label_visibility='collapsed'
        )
    
    with col5:
        # 기본값 설정 (session_state에 없을 때만)
        if f'item_usable_{i}' not in st.session_state:
            st.session_state[f'item_usable_{i}'] = True
        
        item_usable = st.checkbox(
            f'물품{i+1}', 
            label_visibility='collapsed',
            key=f'item_usable_{i}'
        )
        st.write("")
    
    if item_usable and item_price > 0:
        item_names.append(item_name if item_name else '')
        item_prices.append(item_price)
        min_quantities.append(item_min)
        max_quantities.append(item_max)

# 버튼 및 정보 표시
col_left, col_label_fixed, col_right = st.columns([2, 9, 2])

def add_item():
    st.session_state.item_count += 1

with col_left:
    if st.button("물품추가", on_click=add_item, type="primary"):
        pass

with col_label_fixed:
    fixed_budget = sum(a * b for a, b in zip(min_quantities, item_prices))
    max_limit_total = sum(a * b for a, b in zip(max_quantities, item_prices))
    st.warning(
        f"확정: {fixed_budget:,d}원(남은 예산: {(budget_input - fixed_budget):,d}원) 구매제한: {max_limit_total:,d}원"
    )

# 계산 버튼
with col_right:
    if st.button("계산하기", type="primary"):
        if budget_input == "" or budget_input <= 0:
            result_text = '예산을 정확히 입력하세요.(*0보다 큰 자연수)'
        elif len(item_prices) <= 1:
            result_text = '최소 2종류 이상의 단가를 입력하세요.'
        elif min(item_prices) <= 0:
            result_text = '단가가 0보다 작거나 같습니다.'
        elif max(item_prices) > budget_input:
            result_text = '예산이 부족합니다.'
        elif max_limit_total < budget_input:
            result_text = f'최대구매금액({max_limit_total:,d}원)이 예산({budget_input:,d}원)보다 작아 예산을 다 쓸 수 없습니다.'
        elif fixed_budget > budget_input:
            result_text = f'최소구매금액({fixed_budget:,d}원)이 예산({budget_input:,d}원)보다 많아 예산 내에서 쓸 수 없습니다.'
        elif len(item_prices) != len(set(item_prices)):
            result_text = '중복된 단가가 있습니다.'
        else:
            overlay_container = st.empty()
            overlay_container.markdown("""
            <style>
            .overlay {
                position: fixed;top: 0;left: 0;width: 100%;height: 100%;
                background: rgba(0, 0, 0, 0.5);z-index: 999;display: flex;
                justify-content: center;align-items: center;}
            .spinner {margin-bottom: 10px;}
            </style>
            <div class="overlay"><div><div class="spinner">
                        <span class="fa fa-spinner fa-spin fa-3x"></span>
                    </div><div style="color: white;">계산 중...</div></div></div>""", unsafe_allow_html=True)

            result_text, result_list, result_prices, result_labels = calculate_budget(
                budget_input, item_names, item_prices, min_quantities, max_quantities
            )
            overlay_container.empty()

# 결과 출력
if len(result_text.split('\n')) < 30:
    st.code(result_text, language="java")
else:
    st.text_area("결과 출력", result_text, height=300)

# DataFrame 결과 및 다운로드
df_result = None
try:
    df_result = pd.DataFrame(result_list, columns=[f'{price:,d}원' for price in result_prices])
    df_result['금액'] = df_result.mul(result_prices).sum(axis=1)
    
    if len(df_result) > 0:
        st.dataframe(df_result, hide_index=True, use_container_width=True)
        
        # 결과 엑셀 다운로드 버튼 - result_labels 전달
        result_excel = create_result_excel(result_text, df_result, result_labels)
        st.download_button(
            label="📥 결과 다운로드 (Excel)",
            data=result_excel,
            file_name="예산계산_결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
except:
    pass
